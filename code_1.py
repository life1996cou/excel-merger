import os
os.environ['TK_SILENCE_DEPRECATION'] = '1'  # 禁用 macOS 输入法警告

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

def merge_excel():
    try:
        # 获取用户输入
        skip_header = int(entry_header.get())
        skip_footer = int(entry_footer.get())
        folder_path = entry_folder.get()
        width_mode = width_choice.get()

        # 验证文件夹
        if not os.path.isdir(folder_path):
            messagebox.showerror("错误", "文件夹路径无效！")
            return

        merged_wb = None  # 合并后的工作簿
        merged_ws = None  # 合并后的工作表
        current_row = 1   # 当前写入行

        for file in os.listdir(folder_path):
            if file.endswith((".xlsx", ".xls")):
                file_path = os.path.join(folder_path, file)
                wb = load_workbook(file_path)
                ws = wb.active

                # 计算数据区域
                data_start_row = skip_header + 1
                data_end_row = ws.max_row - skip_footer
                data_rows = list(ws.iter_rows(min_row=data_start_row, max_row=data_end_row))

                # 初始化合并工作簿（以第一个文件为模板）
                if merged_wb is None:
                    merged_wb = load_workbook(file_path)
                    merged_ws = merged_wb.active
                    merged_ws.delete_rows(1, merged_ws.max_row)  # 清空数据保留格式
                    current_row = 1
                else:
                    # 追加数据行
                    for row in data_rows:
                        current_row += 1
                        for col_idx, cell in enumerate(row, 1):
                            merged_ws.cell(row=current_row, column=col_idx, value=cell.value)

                    # 复制合并单元格格式
                    for merge_range in ws.merged_cells.ranges:
                        # 计算新合并范围
                        new_min_row = merge_range.min_row - skip_header + (current_row - len(data_rows))
                        new_max_row = merge_range.max_row - skip_header + (current_row - len(data_rows))
                        new_min_col = merge_range.min_col
                        new_max_col = merge_range.max_col

                        # 转换为 Excel 范围格式（如 "A1:B2"）
                        new_range = (
                            f"{get_column_letter(new_min_col)}{new_min_row}:"
                            f"{get_column_letter(new_max_col)}{new_max_row}"
                        )
                        try:
                            merged_ws.merge_cells(new_range)
                        except ValueError as e:
                            print(f"跳过无效合并范围: {new_range}")

                # 记录列宽（模式1）
                if width_mode == 1:
                    for col_idx, col in enumerate(ws.columns, 1):
                        max_length = max(len(str(cell.value)) for cell in col)
                        col_letter = get_column_letter(col_idx)
                        merged_ws.column_dimensions[col_letter].width = max_length + 2

        # 自动调整列宽（模式2）
        if width_mode == 2 and merged_ws:
            for col in merged_ws.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                col_letter = get_column_letter(col[0].column)
                merged_ws.column_dimensions[col_letter].width = max_length + 2

        # 保存文件
        if merged_wb:
            output_path = os.path.join(folder_path, "合并结果.xlsx")
            merged_wb.save(output_path)
            messagebox.showinfo("成功", f"文件已保存为：{output_path}")
        else:
            messagebox.showerror("错误", "文件夹内没有 Excel 文件！")

    except Exception as e:
        messagebox.showerror("错误", f"发生错误：{str(e)}")

# GUI 界面
root = tk.Tk()
root.title("Excel 合并工具")

# 控件样式
style = ttk.Style()
style.configure("TLabel", padding=5)
style.configure("TButton", padding=5)

# 输入控件
ttk.Label(root, text="跳过表头行数:").grid(row=0, column=0, sticky="w")
entry_header = ttk.Entry(root)
entry_header.grid(row=0, column=1, sticky="ew")
entry_header.insert(0, "0")

ttk.Label(root, text="跳过末尾行数:").grid(row=1, column=0, sticky="w")
entry_footer = ttk.Entry(root)
entry_footer.grid(row=1, column=1, sticky="ew")
entry_footer.insert(0, "0")

ttk.Label(root, text="选择文件夹:").grid(row=2, column=0, sticky="w")
entry_folder = ttk.Entry(root, width=40)
entry_folder.grid(row=2, column=1, sticky="ew")

def select_folder():
    folder = filedialog.askdirectory()
    if folder:
        entry_folder.delete(0, tk.END)
        entry_folder.insert(0, folder)

ttk.Button(root, text="浏览...", command=select_folder).grid(row=2, column=2)

# 列宽模式选择
width_choice = tk.IntVar(value=2)
ttk.Label(root, text="列宽模式:").grid(row=3, column=0, sticky="w")
ttk.Radiobutton(root, text="原表格列宽（可能不稳定）", variable=width_choice, value=1).grid(row=3, column=1, sticky="w")
ttk.Radiobutton(root, text="自动调整列宽（推荐）", variable=width_choice, value=2).grid(row=4, column=1, sticky="w")

# 合并按钮
ttk.Button(root, text="开始合并", command=merge_excel).grid(row=5, column=1, pady=10)

# 布局调整
root.columnconfigure(1, weight=1)
root.mainloop()
